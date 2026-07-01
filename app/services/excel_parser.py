from io import BytesIO
from openpyxl import load_workbook
from fastapi import UploadFile

async def parser(file: UploadFile):
    contents = await file.read()
    wb = load_workbook(BytesIO(contents))
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    dados = [dict(zip(headers, row)) for row in rows[1:]]

    return dados
